import csv
import os
import pickle
import random
import re
import string
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from functools import wraps
from io import StringIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, redirect, flash, send_file, jsonify, session, g
from flask_sqlalchemy import SQLAlchemy
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['SECRET_KEY'] = 'mysecretkey'
app.config['UPLOAD_FOLDER'] = r"./uploads"
app.config['PROCESSED_FOLDER'] = r"./processed"
app.config['ALLOWED_EXTENSIONS'] = {'pdf'}
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Ensure folders exist
@app.after_request
def add_header(response):
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    return response

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)
os.makedirs(app.instance_path, exist_ok=True)

MODEL_PATH = os.path.join(app.instance_path, 'ml_model.pkl')

# Database models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False)
    unique_id = db.Column(db.String(10), unique=True, nullable=False)
    uploaded_files = db.relationship('UploadedFile', backref='user', lazy=True, cascade="all, delete-orphan")
    saved_mappings = db.relationship('SavedMapping', backref='user', lazy=True, cascade="all, delete-orphan")

class UploadedFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(120), nullable=False)
    upload_time = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    opening_balance = db.Column(db.Float, default=0.0)
    closing_balance = db.Column(db.Float, default=0.0)
    transactions = db.relationship('Transaction', backref='file', lazy=True, cascade="all, delete-orphan")

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    file_id = db.Column(db.Integer, db.ForeignKey('uploaded_file.id'), nullable=False)
    date = db.Column(db.String(10), nullable=False) # DD-MM-YYYY
    particulars = db.Column(db.String(500), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    withdrawal = db.Column(db.Float, default=0.0)
    deposit = db.Column(db.Float, default=0.0)
    is_manual = db.Column(db.Boolean, default=False)

class SavedMapping(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    merchant = db.Column(db.String(255), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    __table_args__ = (db.UniqueConstraint('user_id', 'merchant', name='_user_merchant_uc'),)

def generate_unique_user_id():
    """Generate a unique random alphanumeric profile ID (e.g., RS-4927)."""
    while True:
        uid = "RS-" + "".join(random.choices(string.digits, k=4))
        # Use modern db.session.scalar to check existence
        exists = db.session.scalar(db.select(User).filter_by(unique_id=uid))
        if not exists:
            return uid

def categorize_transaction(particulars):
    """Categorize transaction using rule-based keywords."""
    p_upper = particulars.upper()
    
    # Medical & Healthcare (Medicines, Hospitals, Doctor consultation, diagnostics)
    if any(k in p_upper for k in [
        'APOLLO', 'PHARMACY', 'MEDICINE', 'HOSPITAL', 'CLINIC', 'MEDPLUS', 'NETMEDS', 
        'PHARMEASY', 'DOCTOR', 'HEALTHCARE', 'DENTAL', 'DIAGNOSTIC', 'LABS', 'SURGERY', 
        'CHEMIST', 'DR.', 'CLINICAL', 'DENTIST', 'OPTICAL', 'LENSKART', 'PHYSIO', 'MEDICOS', 'MEDIC'
    ]):
        return 'Medical & Healthcare'
        
    # Entertainment & Subscriptions
    if any(k in p_upper for k in [
        'NETFLIX', 'SPOTIFY', 'BOOKMYSHOW', 'HOTSTAR', 'YOUTUBE PREMIUM', 'PRIME VIDEO', 
        'PLAYSTATION', 'STEAM', 'GAMING', 'THEATRE', 'CINEMA', 'JIO SAAVN', 'MUSIC',
        'SONYLIV', 'ZEE5', 'APPLE.COM/BILL', 'APPLE SERVICES', 'GOOGLE PLAY', 'PVR', 'INOX',
        'TICKET', 'EVENT', 'DISNEY'
    ]):
        return 'Entertainment & Subscriptions'
        
    # Rent & Maintenance
    if any(k in p_upper for k in ['RENT', 'MAINTENANCE', 'SOCIETY', 'FLAT', 'HOUSE RENT', 'PG PAY', 'LANDLORD', 'HOSTEL', 'NOBROKER', 'MAINT']):
        return 'Rent & Maintenance'
        
    # Education
    if any(k in p_upper for k in [
        'FEES', 'SCHOOL', 'COLLEGE', 'TUITION', 'UDEMY', 'COURSERA', 'BOOKS', 'STATIONERY', 
        'ACADEMY', 'INSTITUTE', 'CLASSES', 'BYJUS', 'UNACADEMY', 'PHYSICS WALLAH', 'SIMPLILEARN', 
        'EDX', 'XEROX', 'PHOTOCOPY'
    ]):
        return 'Education'
    
    # Food & Dining
    if any(k in p_upper for k in [
        'ZOMATO', 'SWIGGY', 'RESTAURANT', 'FOOD', 'CAFE', 'DOMINOS', 'PIZZA', 'EATCLUB', 'DINING',
        'STARBUCKS', 'SUBWAY', 'KFC', 'MCDONALD', 'BURGER KING', 'HALDIRAM', 'BIRYANI', 'BAKERY', 
        'CHAI', 'COFFEE', 'HOTEL', 'DHABA', 'DABA', 'BAR', 'BREWERY', 'PIZZERIA', 'BOX8', 'MOJO PIZZA', 
        'PIZZA HUT', 'TACO BELL', 'EATURE', 'SWEET', 'IRANI CHAI'
    ]):
        return 'Food & Dining'
    
    # Shopping
    if any(k in p_upper for k in [
        'AMAZON', 'FLIPKART', 'MYNTRA', 'RETAIL', 'AJIO', 'SHOPPING', 'SPAR', 'DMART', 'GROCERY', 
        'SUPERMARKET', 'BLINKIT', 'ZEPTO', 'INSTAMART', 'BIGBASKET', 'DUNZO', 'MILKBASKET', 
        'COUNTRY DELIGHT', 'JIO MART', 'JIOMART', 'MEESHO', 'NYKAA', 'TATA CLIQ', 'DECATHLON', 
        'ZARA', 'H&M', 'UNIQLO', 'LIFESTYLE', 'MAX RETAIL', 'TRENDS', 'RELIANCE TRENDS', 'WESTSIDE', 
        'MALL', 'STORE', 'SUPER MARKET', 'MART', 'PROVISIONS', 'APPAREL', 'CLOTHING', 'FOOTWEAR', 
        'JEWELLER', 'KIRANA', 'MILK', 'AMUL', 'MOTHER DAIRY', 'SUPER STORE'
    ]):
        return 'Shopping'
    
    # Salary / Income (Deposits)
    if any(k in p_upper for k in ['SALARY', 'DIVIDEND', 'INTEREST CREDITED', 'INT.PD', 'SCHEME', 'REIMBURSE', 'CASHBACK', 'REFUND']):
        return 'Salary & Income'
    
    # Bills, Utilities & Subscriptions
    if any(k in p_upper for k in [
        'ELECTRICITY', 'TELEPHONE', 'MOBILE', 'RECHARGE', 'BROADBAND', 'INSUR', 'LIC', 'BILLPAY', 
        'GAS', 'DTH', 'WATER BILL', 'BESCOM', 'MSEB', 'MSEDCL', 'UPPCL', 'TNEB', 'BSNL', 'AIRTEL', 
        'JIO', 'VI ', 'VODAFONE', 'IDEA', 'ACT FIBER', 'HATHWAY', 'INSURANCE', 'HDFC ERGO', 
        'ICICI LOMBARD', 'MAX LIFE', 'SBI LIFE', 'FASTAG', 'TOLL', 'WIFI', 'WI-FI', 'POWER', 'DISCOM',
        'TATAPLAY', 'TATA PLAY'
    ]):
        return 'Bills & Utilities'
    
    # Investment
    if any(k in p_upper for k in [
        'MUTUAL FUND', 'INDMONEY', 'ZERODHA', 'GROWW', 'STOCK', 'SIP', 'NPS', 'PPF', 'COIN', 
        'WZRX', 'COINDCX', 'KUBER', 'BINANCE', 'GOLD', 'SECURITIES', 'BROKER', 'UPSTOX', 'ANGELONE', 
        'ANGEL BROKING', '5PAISA', 'FINVASIA', 'MOTILAL', 'SHAREKHAN'
    ]):
        return 'Investment'
        
    # Travel & Fuel
    if any(k in p_upper for k in [
        'UBER', 'OLA', 'PETROLEUM', 'HPCL', 'BPCL', 'IOCL', 'FUEL', 'IRCTC', 'METRO', 'FLIGHT', 
        'TRAVEL', 'RAPIDO', 'NAMMAYATRI', 'NAMMA YATRI', 'BLUSMART', 'MAKEMYTRIP', 'MMT', 'GOIBIBO', 
        'YATRA', 'EASEMYTRIP', 'AIR INDIA', 'INDIGO', 'SPICEJET', 'AKASA', 'RAILWAY', 'BUS', 
        'REDBUS', 'RED BUS', 'KSRTC', 'MSRTC', 'AUTO', 'CAB', 'TAXI', 'SHELL', 'HP FUEL', 'IOC FUEL', 
        'BP FUEL', 'GAS STATION'
    ]):
        return 'Travel & Fuel'
        
    # Cash Withdrawal
    if any(k in p_upper for k in ['ATM', 'CASH WDL', 'CASH WITHDRAWAL', 'WDL', 'WITHDRWL']):
        return 'Cash Withdrawal'

    # Transfers / UPI (Generic)
    # Recognize clean UPI transactions (starts with 'DR / ' or 'CR / ') and general transfer patterns
    if (particulars.startswith('DR / ') or particulars.startswith('CR / ') or 
        any(k in p_upper for k in ['UPI', 'NEFT', 'IMPS', 'RTGS', 'TRANSFER', 'GPAY', 'PAYTM', 'PHONEPE', 'BHIM', 'PAY TO', 'SENT TO', 'RECEIVED FROM'])):
        return 'Transfer'
        
    return 'Other'

def get_merchant_key(particulars):
    """Normalize and extract a merchant key from transaction particulars."""
    parts = [p.strip() for p in particulars.split('/')]
    if len(parts) >= 3 and parts[0] in ('DR', 'CR'):
        # Cleaned UPI format: DR / MERCHANT NAME / BANK NAME / UPI ID
        # Return the merchant name part (uppercased)
        return parts[1].upper()
    return particulars.upper()

def train_classifier():
    """Train the Naive Bayes classifier on current transaction and custom mapping data."""
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.naive_bayes import MultinomialNB
        from sklearn.pipeline import Pipeline

        # Query all categorized transactions (excluding low confidence defaults unless manually reviewed)
        txs = Transaction.query.filter(Transaction.category != 'Other').all()
        # Also query all custom user hard-mappings
        mappings = SavedMapping.query.all()

        texts = []
        labels = []

        # Bootstrap with categorized transactions (using extracted merchant keys)
        for tx in txs:
            texts.append(get_merchant_key(tx.particulars))
            labels.append(tx.category)

        # Boost the custom mappings (give them more sample representation or simple add)
        for m in mappings:
            texts.append(m.merchant)
            labels.append(m.category)

        if len(set(labels)) < 2:
            print("ML_TRAIN: Too few classes (< 2) to train Naive Bayes. Skipping.", flush=True)
            return False

        pipeline = Pipeline([
            ('tfidf', TfidfVectorizer(ngram_range=(1, 2), token_pattern=r'(?u)\b\w+\b')),
            ('nb', MultinomialNB(fit_prior=False))
        ])
        pipeline.fit(texts, labels)

        # Save model to disk
        os.makedirs(os.path.dirname(MODEL_PATH), exist_ok=True)
        with open(MODEL_PATH, 'wb') as f:
            pickle.dump(pipeline, f)
        print(f"ML_TRAIN: Model trained successfully with {len(texts)} samples across {len(set(labels))} classes.", flush=True)
        return True
    except Exception as e:
        print("ML_TRAIN: Error training Naive Bayes classifier:", e, flush=True)
        return False

def train_classifier_in_background():
    """Retrain the classifier in a background thread to prevent blocking requests."""
    def bg_target(app_ctx):
        with app_ctx:
            try:
                train_classifier()
            except Exception as e:
                print("ML_BG: Background classifier training failed:", e, flush=True)

    import threading
    app_ctx = app.app_context()
    thread = threading.Thread(target=bg_target, args=(app_ctx,), daemon=True)
    thread.start()

_classifier = None

def load_classifier():
    global _classifier
    if os.path.exists(MODEL_PATH):
        try:
            with open(MODEL_PATH, 'rb') as f:
                _classifier = pickle.load(f)
        except Exception as e:
            print("ML_LOAD: Error loading classifier pickle:", e, flush=True)
            _classifier = None

def predict_category(particulars):
    """Predict category for transaction particulars using the ML model.
    Returns (predicted_category, confidence) or (None, 0.0).
    """
    global _classifier
    if _classifier is None:
        load_classifier()
    if _classifier is None:
        return None, 0.0

    try:
        # Predict probability using extracted merchant key
        merchant_key = get_merchant_key(particulars)
        probs = _classifier.predict_proba([merchant_key])[0]
        classes = _classifier.classes_
        max_idx = probs.argmax()
        confidence = float(probs[max_idx])
        category = classes[max_idx]
        return category, confidence
    except Exception as e:
        print("ML_PREDICT: Error in Naive Bayes prediction:", e, flush=True)
        return None, 0.0

def extract_balances(text):
    """Extract opening and closing balances from PDF text."""
    opening_bal = 0.0
    closing_bal = 0.0
    
    lines = text.split('\n')
    
    # 1. Search for Opening Balance
    for idx, line in enumerate(lines):
        line_clean = line.strip()
        match = re.search(r'Your Opening Balance on \d{2}-\d{2}-\d{2,4}:\s*([\d,]+\.\d{2})', line_clean, re.IGNORECASE)
        if match:
            opening_bal = float(match.group(1).replace(',', ''))
            break
        elif re.search(r'Your Opening Balance on \d{2}-\d{2}-\d{2,4}:', line_clean, re.IGNORECASE):
            found = False
            for offset in range(1, 12):
                if idx + offset < len(lines):
                    sub_line = lines[idx + offset].strip()
                    val_match = re.search(r'(?:Rs\.?|INR)?\s*([\d,]+\.\d{2})\s*(?:Cr|Dr)?', sub_line, re.IGNORECASE)
                    if val_match:
                        try:
                            opening_bal = float(val_match.group(1).replace(',', ''))
                            found = True
                            break
                        except ValueError:
                            continue
            if found:
                break

    # 2. Search for Closing Balance
    for idx, line in enumerate(lines):
        line_clean = line.strip()
        match = re.search(r'Your Closing Balance on \d{2}-\d{2}-\d{2,4}:\s*([\d,]+\.\d{2})', line_clean, re.IGNORECASE)
        if match:
            closing_bal = float(match.group(1).replace(',', ''))
            break
        elif re.search(r'Your Closing Balance on \d{2}-\d{2}-\d{2,4}:', line_clean, re.IGNORECASE):
            found = False
            for offset in range(1, 12):
                if idx + offset < len(lines):
                    sub_line = lines[idx + offset].strip()
                    val_match = re.search(r'(?:Rs\.?|INR)?\s*([\d,]+\.\d{2})\s*(?:Cr|Dr)?', sub_line, re.IGNORECASE)
                    if val_match:
                        try:
                            closing_bal = float(val_match.group(1).replace(',', ''))
                            found = True
                            break
                        except ValueError:
                            continue
            if found:
                break
                
    return opening_bal, closing_bal

# Ensure database tables are created on the first request
first_request_executed = False

@app.before_request
def create_tables():
    global first_request_executed
    if not first_request_executed:
        db.create_all()
        
        # Add opening_balance and closing_balance columns if not present
        try:
            db.session.execute(db.text("ALTER TABLE uploaded_file ADD COLUMN opening_balance FLOAT DEFAULT 0.0"))
            db.session.execute(db.text("ALTER TABLE uploaded_file ADD COLUMN closing_balance FLOAT DEFAULT 0.0"))
            db.session.commit()
            print("MIGRATION: Added opening_balance and closing_balance columns to uploaded_file table.", flush=True)
        except Exception:
            db.session.rollback()

        # Add is_manual column to transaction if not present
        try:
            db.session.execute(db.text('ALTER TABLE "transaction" ADD COLUMN is_manual BOOLEAN DEFAULT 0'))
            db.session.commit()
            print("MIGRATION: Added is_manual column to transaction table.", flush=True)
        except Exception:
            db.session.rollback()
            
        # Run one-time migration to upgrade existing 'Other' transactions
        try:
            txs = Transaction.query.filter_by(category='Other').all()
            print(f"MIGRATION: Found {len(txs)} transactions with 'Other' category.", flush=True)
            updated = False
            for tx in txs:
                new_cat = categorize_transaction(tx.particulars)
                if new_cat != 'Other':
                    print(f"MIGRATION: Updating '{tx.particulars}' from 'Other' -> '{new_cat}'", flush=True)
                    tx.category = new_cat
                    updated = True
            if updated:
                db.session.commit()
                print("MIGRATION: Database committed successfully.", flush=True)
        except Exception as e:
            print("Database migration error (Other category):", e, flush=True)
            
        # Populate opening and closing balances for existing statements
        try:
            files = UploadedFile.query.all()
            updated_files = False
            for f in files:
                if (f.opening_balance is None or f.opening_balance == 0.0) and (f.closing_balance is None or f.closing_balance == 0.0):
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
                    if os.path.exists(file_path):
                        rsrcmgr = PDFResourceManager()
                        retstr = StringIO()
                        codec = 'utf-8'
                        laparams = LAParams(all_texts=True, detect_vertical=False, char_margin=50.0)
                        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
                        fp = open(file_path, 'rb')
                        interpreter = PDFPageInterpreter(rsrcmgr, device)
                        for page in PDFPage.get_pages(fp, caching=False):
                            interpreter.process_page(page)
                        text = retstr.getvalue()
                        fp.close()
                        device.close()
                        retstr.close()
                        
                        op_bal, cl_bal = extract_balances(text)
                        f.opening_balance = op_bal
                        f.closing_balance = cl_bal
                        updated_files = True
                        print(f"MIGRATION: Populated balance for {f.filename}: Op={op_bal}, Cl={cl_bal}", flush=True)
            if updated_files:
                db.session.commit()
                print("MIGRATION: Saved balances to database.", flush=True)
        except Exception as e:
            print("Migration error populating balances:", e, flush=True)
            
        # Bootstrap ML model training
        try:
            train_classifier()
        except Exception as e:
            print("MIGRATION: Error bootstrapping Naive Bayes model:", e, flush=True)

        first_request_executed = True

@app.before_request
def load_logged_in_user():
    user_id = session.get('user_id')
    if user_id is None:
        g.user = None
    else:
        g.user = db.session.get(User, user_id)

def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            if request.path.startswith('/api/') or request.headers.get('Accept') == 'application/json':
                return jsonify({'success': False, 'error': 'Unauthorized'}), 401
            return redirect('/login')
        return view(**kwargs)
    return wrapped_view


# Allowed file check
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def clean_amount(amount):
    """Clean and round monetary amounts."""
    cleaned_amount = Decimal(amount.replace(',', ''))
    rounded_amount = cleaned_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return f"{rounded_amount:,.2f}"

def parse_date(date_str):
    """Convert date from DD-MM-YY to DD-MM-YYYY format."""
    return datetime.strptime(date_str, '%d-%m-%y').strftime('%d-%m-%Y')

def clean_particulars(particulars):
    """Clean and structure transaction particulars."""
    # Check if it follows the UPI pattern: UPI/DR or CR/REF/NAME/BANK/UPI_ID/...
    upi_match = re.search(r'UPI/(DR|CR)/(\d+|)//?([^/]+)/([^/]+)/([^/]+)', particulars, re.IGNORECASE)
    if upi_match:
        direction = upi_match.group(1).upper()
        name = upi_match.group(3).strip()
        bank = upi_match.group(4).strip()
        upi_id = upi_match.group(5).strip()
        
        # Clean any trailing dashes/slashes/dots from the fields
        name = re.sub(r'[\s\-.]+$', '', name)
        bank = re.sub(r'[\s\-.]+$', '', bank)
        upi_id = re.sub(r'[\s\-.]+$', '', upi_id)
        
        parts = [direction, name, bank]
        if upi_id:
            parts.append(upi_id)
        return " / ".join(parts)
        
    # Fallback cleaning for non-UPI transactions
    cleaned = re.sub(r'[^\w\s:/\-@#&*(),.]+', '', particulars)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    # Remove trailing noise
    cleaned = re.sub(r'[\s\-.]+$', '', cleaned)
    return cleaned if cleaned else "Unclassified Transaction"

def is_withdrawal(particulars):
    """Determine if a transaction is a withdrawal based on keywords in particulars."""
    p_upper = particulars.upper()
    
    # 1. Direct UPI checks (highly specific)
    if 'UPI/CR/' in p_upper:
        return False
    if 'UPI/DR/' in p_upper:
        return True
        
    # 2. Standalone keywords check
    withdrawal_keywords = ['CHQ RETN', 'TM', 'TRANSFER:', 'RETN', 'INWARD']
    if any(keyword in p_upper for keyword in withdrawal_keywords):
        return True
        
    # 3. Check for Debit marking (DR/Dr) or ACH Debit (ACHDr/ACH DR) using word boundaries
    if re.search(r'\bDR\b|\bACHDR\b', p_upper):
        return True
        
    return False

def process_transaction(date, particulars, grouped_data):
    """Process a single transaction and add it to the grouped data."""
    amounts = re.findall(r'\d{1,5}(?:,\d{3})*\.\d{2}', particulars)
    if amounts:
        amount = clean_amount(amounts[0])
        withdrawal = amount if is_withdrawal(particulars) else '0.00'
        deposit = '0.00' if is_withdrawal(particulars) else amount
        if date not in grouped_data:
            grouped_data[date] = []  # Initialize list for this date
        grouped_data[date].append([date, clean_particulars(particulars), withdrawal, deposit])

def convert_pdf_to_txt(path, csv_path):
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'

    laparams = LAParams(
        all_texts=True, detect_vertical=False,
        line_overlap=0.5, char_margin=50.0,
        line_margin=2.0, word_margin=2,
        boxes_flow=1
    )
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = open(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0  # Process all pages
    caching = False
    pagenos = set()
    grouped_data = {}

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password, caching=caching, check_extractable=True):
        interpreter.process_page(page)

    text = retstr.getvalue()
    lines = text.split('\n')
    particulars = ""
    date = None
    transaction_start = False

    for line in lines:
        line = line.strip()
        if not line or "Page Total:" in line or "----------" in line:
            continue

        date_match = re.match(r'^(\d{2}-\d{2}-(\d{2}))', line)
        if date_match:
            if transaction_start:
                process_transaction(date, particulars, grouped_data)
                particulars = ""
            transaction_start = True
            full_date = date_match.group(1)
            year_suffix = date_match.group(2)
            date = parse_date(full_date)
            particulars = line.split('-' + year_suffix, 1)[-1].strip()
        else:
            particulars += " " + line.strip()

    if transaction_start:
        process_transaction(date, particulars, grouped_data)

    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['DATE', 'PARTICULARS', 'WITHDRAWALS', 'DEPOSITS', 'TRANSACTION COUNT'])
        for date_key, transactions in grouped_data.items():
            # Calculate totals for the day
            total_withdrawal = sum(Decimal(txn[2].replace(',', '')) for txn in transactions if txn[2] != '0.00')
            total_deposit = sum(Decimal(txn[3].replace(',', '')) for txn in transactions if txn[3] != '0.00')
            
            writer.writerow([f"Date: {date_key}", "", "", "", f"Transactions: {len(transactions)}"])
            writer.writerow(['DATE', 'PARTICULARS', 'WITHDRAWALS', 'DEPOSITS'])
            for transaction in transactions:
                writer.writerow(transaction)
            
            # Write totals for the day
            writer.writerow(["", "Total for the day:", f"{total_withdrawal:,.2f}", f"{total_deposit:,.2f}"])
            writer.writerow([])  # Blank row to separate tables
            writer.writerow([])  # Extra blank row 1
            writer.writerow([])  # Extra blank row 2
            writer.writerow([])  # Extra blank row 3

    fp.close()
    device.close()
    retstr.close()
    return text, grouped_data

def generate_excel_report(file_record, excel_path):
    wb = Workbook()
    
    # ------------------ SHEET 1: TRANSACTIONS ------------------
    ws_tx = wb.active
    ws_tx.title = "Transactions"
    ws_tx.views.sheetView[0].showGridLines = True
    
    # Fonts
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    # Fills
    fill_title = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_summary_header = PatternFill(start_color="D1E2F2", end_color="D1E2F2", fill_type="solid")
    
    # Borders
    thin_border_side = Side(style='thin', color='CBD5E1')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=thin_border_side, bottom=Side(style='double', color='1F4E79'))
    
    # Alignments
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # 1. Title Block
    ws_tx.merge_cells("A1:F2")
    title_cell = ws_tx["A1"]
    title_cell.value = "ParsePAY Financial Statement Report"
    title_cell.font = font_title
    title_cell.alignment = align_center
    
    for r in range(1, 3):
        for c in range(1, 7):
            cell = ws_tx.cell(row=r, column=c)
            cell.fill = fill_title
            
    # 2. Metadata Cards (Summary Block)
    ws_tx.merge_cells("A4:C4")
    ws_tx["A4"].value = "STATEMENT METADATA"
    ws_tx["A4"].font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    ws_tx["A4"].fill = fill_summary_header
    ws_tx["A4"].alignment = align_center
    
    ws_tx["A5"].value = "Statement Name:"
    ws_tx["B5"].value = file_record.filename
    ws_tx["A6"].value = "Opening Balance:"
    ws_tx["B6"].value = file_record.opening_balance
    ws_tx["B6"].number_format = '₹#,##0.00'
    ws_tx["A7"].value = "Closing Balance:"
    ws_tx["B7"].value = file_record.closing_balance
    ws_tx["B7"].number_format = '₹#,##0.00'
    
    ws_tx.merge_cells("E4:F4")
    ws_tx["E4"].value = "RECONCILIATION SUMMARY"
    ws_tx["E4"].font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    ws_tx["E4"].fill = fill_summary_header
    ws_tx["E4"].alignment = align_center
    
    transactions = Transaction.query.filter_by(file_id=file_record.id).order_by(Transaction.id).all()
    total_withdrawals = sum(tx.withdrawal for tx in transactions)
    total_deposits = sum(tx.deposit for tx in transactions)
    net_savings = total_deposits - total_withdrawals
    
    ws_tx["E5"].value = "Total Deposits (Cr):"
    ws_tx["F5"].value = total_deposits
    ws_tx["F5"].number_format = '₹#,##0.00'
    ws_tx["F5"].font = font_bold
    
    ws_tx["E6"].value = "Total Withdrawals (Dr):"
    ws_tx["F6"].value = total_withdrawals
    ws_tx["F6"].number_format = '₹#,##0.00'
    ws_tx["F6"].font = font_bold
    
    ws_tx["E7"].value = "Net Flow:"
    ws_tx["F7"].value = net_savings
    ws_tx["F7"].number_format = '₹#,##0.00'
    ws_tx["F7"].font = font_bold
    if net_savings >= 0:
        ws_tx["F7"].font = Font(name="Calibri", size=11, bold=True, color="10B981")
    else:
        ws_tx["F7"].font = Font(name="Calibri", size=11, bold=True, color="EF4444")
        
    for r in range(4, 8):
        for c in [1, 2, 3, 5, 6]:
            cell = ws_tx.cell(row=r, column=c)
            if r > 4:
                if c in (1, 5):
                    cell.font = font_bold
                else:
                    if cell.font != font_bold and cell.font.color is None:
                        cell.font = font_regular
            cell.border = thin_border
            if r == 4:
                cell.fill = fill_summary_header
                
    # 3. Main Data Table
    start_row = 9
    headers = ["S.No.", "Date", "Particulars", "Category", "Withdrawals (Dr)", "Deposits (Cr)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_tx.cell(row=start_row, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in (1, 2, 4) else (align_right if col_idx in (5, 6) else align_left)
        cell.border = thin_border
        
    current_row = start_row + 1
    for idx, tx in enumerate(transactions, 1):
        ws_tx.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws_tx.cell(row=current_row, column=2, value=tx.date).alignment = align_center
        ws_tx.cell(row=current_row, column=3, value=tx.particulars).alignment = align_left
        ws_tx.cell(row=current_row, column=4, value=tx.category).alignment = align_center
        
        w_cell = ws_tx.cell(row=current_row, column=5, value=tx.withdrawal)
        w_cell.number_format = '₹#,##0.00'
        w_cell.alignment = align_right
        
        d_cell = ws_tx.cell(row=current_row, column=6, value=tx.deposit)
        d_cell.number_format = '₹#,##0.00'
        d_cell.alignment = align_right
        
        row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(1, 7):
            cell = ws_tx.cell(row=current_row, column=col_idx)
            cell.font = font_regular
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = thin_border
            
        current_row += 1
        
    # Totals Row
    total_row = current_row
    ws_tx.cell(row=total_row, column=2, value="Total").font = font_bold
    ws_tx.cell(row=total_row, column=2).alignment = align_center
    
    w_sum = ws_tx.cell(row=total_row, column=5, value=f"=SUM(E{start_row+1}:E{total_row-1})")
    w_sum.font = font_bold
    w_sum.number_format = '₹#,##0.00'
    w_sum.alignment = align_right
    w_sum.border = double_bottom_border
    
    d_sum = ws_tx.cell(row=total_row, column=6, value=f"=SUM(F{start_row+1}:F{total_row-1})")
    d_sum.font = font_bold
    d_sum.number_format = '₹#,##0.00'
    d_sum.alignment = align_right
    d_sum.border = double_bottom_border
    
    for c in (1, 3, 4):
        ws_tx.cell(row=total_row, column=c).border = double_bottom_border
        
    for col in ws_tx.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2, 4):
                continue
            val_str = str(cell.value or '')
            if cell.number_format and ('₹' in cell.number_format or '#' in cell.number_format) and isinstance(cell.value, (int, float)):
                val_str = f"Rs. {cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_tx.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws_tx.column_dimensions['C'].width = max(ws_tx.column_dimensions['C'].width, 35)
    
    # ------------------ SHEET 2: CATEGORY BREAKDOWN ------------------
    ws_cat = wb.create_sheet(title="Category Summary")
    ws_cat.views.sheetView[0].showGridLines = True
    
    ws_cat.merge_cells("A1:D2")
    cat_title = ws_cat["A1"]
    cat_title.value = "Category Spending Analysis"
    cat_title.font = font_title
    cat_title.alignment = align_center
    
    for r in range(1, 3):
        for c in range(1, 5):
            ws_cat.cell(row=r, column=c).fill = fill_title
            
    cat_headers = ["Category", "Total Spent (Dr)", "Transaction Count", "% of Total Spent"]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in (3, 4) else (align_right if col_idx == 2 else align_left)
        cell.border = thin_border
        
    category_data = {}
    for tx in transactions:
        if tx.category not in category_data:
            category_data[tx.category] = {"spent": 0.0, "count": 0}
        category_data[tx.category]["count"] += 1
        if tx.withdrawal > 0:
            category_data[tx.category]["spent"] += tx.withdrawal
            
    sorted_categories = sorted(category_data.items(), key=lambda x: x[1]["spent"], reverse=True)
    
    row_idx = 5
    for idx, (cat_name, cat_info) in enumerate(sorted_categories, 1):
        ws_cat.cell(row=row_idx, column=1, value=cat_name).alignment = align_left
        
        s_cell = ws_cat.cell(row=row_idx, column=2, value=cat_info["spent"])
        s_cell.number_format = '₹#,##0.00'
        s_cell.alignment = align_right
        
        ws_cat.cell(row=row_idx, column=3, value=cat_info["count"]).alignment = align_center
        
        p_cell = ws_cat.cell(row=row_idx, column=4, value=f"=B{row_idx}/SUM(B5:B{4+len(sorted_categories)})")
        p_cell.number_format = '0.0%'
        p_cell.alignment = align_center
        
        row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(1, 5):
            cell = ws_cat.cell(row=row_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if row_fill.fill_type:
                cell.fill = row_fill
                
        row_idx += 1
        
    cat_total_row = row_idx
    ws_cat.cell(row=cat_total_row, column=1, value="Total Spending").font = font_bold
    ws_cat.cell(row=cat_total_row, column=1).alignment = align_left
    
    tot_spent = ws_cat.cell(row=cat_total_row, column=2, value=f"=SUM(B5:B{cat_total_row-1})")
    tot_spent.font = font_bold
    tot_spent.number_format = '₹#,##0.00'
    tot_spent.alignment = align_right
    tot_spent.border = double_bottom_border
    
    tot_count = ws_cat.cell(row=cat_total_row, column=3, value=f"=SUM(C5:C{cat_total_row-1})")
    tot_count.font = font_bold
    tot_count.alignment = align_center
    tot_count.border = double_bottom_border
    
    tot_pct = ws_cat.cell(row=cat_total_row, column=4, value=f"=SUM(D5:D{cat_total_row-1})")
    tot_pct.font = font_bold
    tot_pct.number_format = '0.0%'
    tot_pct.alignment = align_center
    tot_pct.border = double_bottom_border
    
    ws_cat.cell(row=cat_total_row, column=1).border = double_bottom_border
    
    for col in ws_cat.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2):
                continue
            val_str = str(cell.value or '')
            if cell.number_format and '%' in cell.number_format:
                val_str = "100.0%"
            elif cell.number_format and '₹' in cell.number_format and isinstance(cell.value, (int, float)):
                val_str = f"Rs. {cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_cat.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    wb.save(excel_path)

@app.route('/api/auth/me')
def auth_me():
    if g.user is None:
        return jsonify({'logged_in': False})
    return jsonify({
        'logged_in': True,
        'user': {
            'id': g.user.id,
            'username': g.user.username,
            'unique_id': g.user.unique_id
        }
    })

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    unique_id = data.get('user_id', '').strip().upper()
    if not unique_id:
        return jsonify({'success': False, 'error': 'User ID cannot be empty'}), 400
        
    user = User.query.filter_by(unique_id=unique_id).first()
    if user:
        session['user_id'] = user.id
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'unique_id': user.unique_id
            }
        })
    return jsonify({'success': False, 'error': 'Invalid User ID. Please check and try again.'}), 404

@app.route('/api/auth/register', methods=['POST'])
def api_register():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    if not username:
        return jsonify({'success': False, 'error': 'Username cannot be empty'}), 400
        
    unique_id = generate_unique_user_id()
    new_user = User(username=username, unique_id=unique_id)
    db.session.add(new_user)
    db.session.commit()
    
    session['user_id'] = new_user.id
    return jsonify({
        'success': True,
        'user': {
            'id': new_user.id,
            'username': new_user.username,
            'unique_id': new_user.unique_id
        }
    })

@app.route('/api/auth/logout', methods=['POST', 'GET'])
def api_logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/statements')
@login_required
def api_statements():
    recent_files = UploadedFile.query.filter_by(user_id=g.user.id).order_by(UploadedFile.upload_time.desc()).all()
    files_list = [{
        'id': rf.id,
        'filename': rf.filename,
        'upload_time': rf.upload_time.strftime('%Y-%m-%d %H:%M:%S') if rf.upload_time else '',
        'opening_balance': rf.opening_balance,
        'closing_balance': rf.closing_balance
    } for rf in recent_files]
    return jsonify({'success': True, 'statements': files_list})


@app.route('/api/upload', methods=['POST'])
@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    is_api = request.path.startswith('/api/') or request.headers.get('Accept') == 'application/json'

    if 'file' not in request.files:
        if is_api:
            return jsonify({'success': False, 'error': 'No file part in the request'}), 400
        flash('No file selected', 'error')
        return redirect('/')

    file = request.files['file']
    if file.filename == '':
        if is_api:
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        flash('No file selected', 'error')
        return redirect('/')

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        csv_filename = f"{filename}.csv"
        csv_path = os.path.join(app.config['PROCESSED_FOLDER'], csv_filename)
        pdf_text, grouped_data = convert_pdf_to_txt(file_path, csv_path)
        op_bal, cl_bal = extract_balances(pdf_text)
        
        # Save statement metadata to database
        db_file = UploadedFile(
            filename=filename, 
            user_id=g.user.id,
            opening_balance=op_bal,
            closing_balance=cl_bal
        )
        db.session.add(db_file)
        db.session.commit()

        # Query all custom merchant mappings for the user at once to avoid N+1 queries
        user_mappings = SavedMapping.query.filter_by(user_id=g.user.id).all()
        mappings_dict = {m.merchant.upper(): m.category for m in user_mappings}

        # Save individual transactions to database
        for date_key, transactions in grouped_data.items():
            for tx in transactions:
                particulars = tx[1]
                try:
                    w_val = float(tx[2].replace(',', ''))
                except ValueError:
                    w_val = 0.0
                try:
                    d_val = float(tx[3].replace(',', ''))
                except ValueError:
                    d_val = 0.0
                    
                # Hierarchical Categorization Pipeline
                merchant_key = get_merchant_key(particulars)
                category = mappings_dict.get(merchant_key)
                
                if not category:
                    pred_cat, confidence = predict_category(particulars)
                    if pred_cat and confidence >= 0.70:
                        category = pred_cat
                    else:
                        category = categorize_transaction(particulars)
                
                db_tx = Transaction(
                    file_id=db_file.id,
                    date=tx[0],
                    particulars=particulars,
                    category=category,
                    withdrawal=w_val,
                    deposit=d_val,
                    is_manual=False
                )
                db.session.add(db_tx)
        db.session.commit()
        
        # Trigger retraining with new bootstrapped or hard-mapped transactions in background
        try:
            train_classifier_in_background()
        except Exception as e:
            print("ML_UPLOAD: Error launching model training after upload:", e, flush=True)

        excel_filename = f"{filename}.xlsx"
        excel_path = os.path.join(app.config['PROCESSED_FOLDER'], excel_filename)
        try:
            generate_excel_report(db_file, excel_path)
            print(f"EXCEL_GEN: Created initial Excel report for {filename}", flush=True)
        except Exception as e:
            print(f"EXCEL_GEN: Error creating initial Excel: {e}", flush=True)
            df = pd.read_csv(csv_path)
            df.to_excel(excel_path, index=None, header=True)

        if is_api:
            return jsonify({'success': True, 'file_id': db_file.id})
        flash('File processed successfully!', 'success')
        return redirect(f'/dashboard/{db_file.id}')

    if is_api:
        return jsonify({'success': False, 'error': 'Invalid file format. Only PDF statements are allowed.'}), 400
    flash('Invalid file format', 'error')
    return redirect('/')


@app.route('/api/dashboard-data/<int:file_id>')
@login_required
def api_dashboard_data(file_id):
    file = UploadedFile.query.get_or_404(file_id)
    if file.user_id != g.user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    transactions = Transaction.query.filter_by(file_id=file_id).all()
    
    total_withdrawals = 0.0
    total_deposits = 0.0
    category_totals = {}
    category_counts = {}
    daily_trends = {}
    
    tx_list = []
    for tx in transactions:
        total_withdrawals += tx.withdrawal
        total_deposits += tx.deposit
        
        # Spent analysis (Withdrawals only)
        if tx.withdrawal > 0:
            category_totals[tx.category] = category_totals.get(tx.category, 0.0) + tx.withdrawal
            
        # Count transactions per category
        category_counts[tx.category] = category_counts.get(tx.category, 0) + 1
        
        # Daily trends
        if tx.date not in daily_trends:
            daily_trends[tx.date] = {'withdrawal': 0.0, 'deposit': 0.0}
        daily_trends[tx.date]['withdrawal'] += tx.withdrawal
        daily_trends[tx.date]['deposit'] += tx.deposit
        
        tx_list.append({
            'id': tx.id,
            'date': tx.date,
            'particulars': tx.particulars,
            'category': tx.category,
            'withdrawal': tx.withdrawal,
            'deposit': tx.deposit
        })
        
    # Sort trends chronologically
    sorted_dates = sorted(daily_trends.keys(), key=lambda d: datetime.strptime(d, '%d-%m-%Y'))
    trend_labels = sorted_dates
    trend_withdrawals = [daily_trends[d]['withdrawal'] for d in sorted_dates]
    trend_deposits = [daily_trends[d]['deposit'] for d in sorted_dates]
    
    return jsonify({
        'filename': file.filename,
        'upload_time': file.upload_time.strftime('%Y-%m-%d %H:%M:%S') if file.upload_time else '',
        'total_withdrawals': round(total_withdrawals, 2),
        'total_deposits': round(total_deposits, 2),
        'net_savings': round(total_deposits - total_withdrawals, 2),
        'transaction_count': len(transactions),
        'opening_balance': round(file.opening_balance, 2) if file.opening_balance is not None else 0.0,
        'closing_balance': round(file.closing_balance, 2) if file.closing_balance is not None else 0.0,
        'category_totals': {k: round(v, 2) for k, v in category_totals.items()},
        'category_counts': category_counts,
        'daily_trends': {
            'labels': trend_labels,
            'withdrawals': trend_withdrawals,
            'deposits': trend_deposits
        },
        'transactions': tx_list
    })

@app.route('/api/update-category', methods=['POST'])
@login_required
def update_category():
    data = request.get_json() or {}
    tx_id = data.get('transaction_id')
    new_category = data.get('category')
    
    if not tx_id or not new_category:
        return jsonify({'success': False, 'error': 'Missing parameters'}), 400
        
    tx = Transaction.query.get(tx_id)
    if not tx:
        return jsonify({'success': False, 'error': 'Transaction not found'}), 404
        
    file = UploadedFile.query.get(tx.file_id)
    if file.user_id != g.user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
    tx.category = new_category
    tx.is_manual = True
    
    # Upsert custom hard mapping for this merchant key under user profile
    merchant_key = get_merchant_key(tx.particulars)
    mapping = SavedMapping.query.filter_by(user_id=g.user.id, merchant=merchant_key).first()
    if mapping:
        mapping.category = new_category
    else:
        mapping = SavedMapping(user_id=g.user.id, merchant=merchant_key, category=new_category)
        db.session.add(mapping)
        
    db.session.commit()
    
    # Retrain ML model with new data in background
    try:
        train_classifier_in_background()
    except Exception as e:
        print("ML_ROUTE: Error launching background model training after manual update:", e, flush=True)
        
    return jsonify({'success': True})

@app.route('/api/swap-transaction-amount', methods=['POST'])
@login_required
def swap_transaction_amount():
    data = request.get_json() or {}
    tx_id = data.get('transaction_id')
    
    if not tx_id:
        return jsonify({'success': False, 'error': 'Missing transaction_id'}), 400
        
    tx = Transaction.query.get(tx_id)
    if not tx:
        return jsonify({'success': False, 'error': 'Transaction not found'}), 404
        
    file = UploadedFile.query.get(tx.file_id)
    if file.user_id != g.user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
    # Swap withdrawal and deposit
    tx.withdrawal, tx.deposit = tx.deposit, tx.withdrawal
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/delete-statement/<int:file_id>', methods=['POST', 'DELETE'])
@app.route('/delete-statement/<int:file_id>', methods=['POST'])
@login_required
def delete_statement(file_id):
    is_api = request.path.startswith('/api/') or request.headers.get('Accept') == 'application/json'
    file = UploadedFile.query.get_or_404(file_id)
    if file.user_id != g.user.id:
        if is_api:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        flash('Unauthorized action.', 'error')
        return redirect('/')
    
    # Delete local files
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    if os.path.exists(upload_path):
        try:
            os.remove(upload_path)
        except Exception:
            pass
            
    csv_path = os.path.join(app.config['PROCESSED_FOLDER'], f"{file.filename}.csv")
    if os.path.exists(csv_path):
        try:
            os.remove(csv_path)
        except Exception:
            pass
            
    excel_path = os.path.join(app.config['PROCESSED_FOLDER'], f"{file.filename}.xlsx")
    if os.path.exists(excel_path):
        try:
            os.remove(excel_path)
        except Exception:
            pass
            
    db.session.delete(file)
    db.session.commit()
    
    if is_api:
        return jsonify({'success': True})
    flash('Statement deleted successfully!', 'success')
    return redirect('/')


@app.route('/download/<filename>')
@login_required
def download_file(filename):
    # Retrieve file entry to verify ownership
    # The filename matches the output excel, which is secure_filename(pdf_name) + ".xlsx"
    orig_pdf_name = filename.rsplit('.', 1)[0]
    file_record = UploadedFile.query.filter_by(filename=orig_pdf_name, user_id=g.user.id).first()
    if not file_record:
        # Fallback search inside the user's files
        user_files = UploadedFile.query.filter_by(user_id=g.user.id).all()
        for uf in user_files:
            if f"{uf.filename}.xlsx" == filename:
                file_record = uf
                break
        if not file_record:
            flash('Access to this download is restricted.', 'error')
            return redirect('/')
            
    file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
    
    # Dynamically regenerate Excel file before serving to include latest category edits
    try:
        generate_excel_report(file_record, file_path)
        print(f"EXCEL_GEN: Dynamically regenerated Excel report for {filename}", flush=True)
    except Exception as e:
        print(f"EXCEL_GEN: Error dynamically regenerating Excel for {filename}: {e}", flush=True)
        
    return send_file(file_path, as_attachment=True)

@app.route('/debug-text/<int:file_id>')
@login_required
def debug_text(file_id):
    file = UploadedFile.query.get_or_404(file_id)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams(all_texts=True, detect_vertical=False, char_margin=50.0)
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = open(file_path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    for page in PDFPage.get_pages(fp, caching=False):
        interpreter.process_page(page)
    text = retstr.getvalue()
    fp.close()
    device.close()
    retstr.close()
    
    lines = text.split('\n')
    output_lines = []
    
    # 1. Print lines around opening balance (known to be around 288)
    output_lines.append("=== LINES 270-320 ===")
    for idx in range(max(0, 270), min(len(lines), 320)):
        output_lines.append(f"{idx}: {lines[idx]}")
        
    # 2. Print lines around closing balance
    output_lines.append("\n=== LINES WITH CLOSING BALANCE AND SURROUNDINGS ===")
    for idx, line in enumerate(lines):
        if "closing balance" in line.lower():
            output_lines.append(f"\nFound closing balance at line {idx}:")
            for offset in range(-5, 10):
                target_idx = idx + offset
                if 0 <= target_idx < len(lines):
                    output_lines.append(f"{target_idx}: {lines[target_idx]}")
                    
    # 3. Print lines containing 'balance'
    output_lines.append("\n=== ALL LINES WITH 'BALANCE' ===")
    for idx, line in enumerate(lines):
        if "balance" in line.lower():
            output_lines.append(f"{idx}: {line}")
            
    return "<pre>" + "\n".join(output_lines) + "</pre>"

from flask import send_from_directory, abort

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_spa(path):
    if path.startswith('api/') or path.startswith('upload') or path.startswith('download/') or path.startswith('delete-statement/') or path.startswith('debug-text/'):
        abort(404)
    if os.path.exists(os.path.join(app.static_folder, 'index.html')):
        return send_from_directory(app.static_folder, 'index.html')
    return "React SPA build files not found in the 'static' folder. Please build the frontend using 'npm run build' inside 'frontend' directory or access via Vite dev server at http://localhost:5173"

if __name__ == '__main__':
    app.run(debug=True, port=5000)

